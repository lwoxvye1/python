# 读写csv文件
import csv


def readCsv(path):
    infoList = []
    with open(path, "r") as f:
        allFileInfo = csv.reader(f)
        for row in allFileInfo:
            infoList.append(row)
    return infoList


def writeCsv(path, data):
    with open(path, "w") as f:
        writer = csv.writer(f)
        for rowData in data:
            writer.writerow(rowData)


path = "D:\学习\python\project\csv.csv"
writeCsv(path, [[1, 2, 3], [4, 5, 6], [7, 8, 9]])
infoList = readCsv(path)
print(infoList)


# 读取PDF文件
import sys
import importlib
importlib.reload(sys)

from pdfminer.pdfinterp import PDFPageInterpreter, PDFResourceManager
from pdfminer.pdfparser import PDFDocument, PDFParser
from pdfminer.layout import LAParams, LTTextBoxHorizontal
from pdfminer.converter import PDFPageAggregator
from pdfminer.pdfinterp import PDFTextExtractionNotAllowed


def readPDF(path, callback=None, topath=""):
    # 以二进制形式打开pdf文件
    f = open(path, "rb")
    # 创建一个pdf文档分析器
    parser = PDFParser(f)
    # 创建pdf文档
    pdfFile = PDFDocument()
    # 链接分析器与文档对象
    parser.set_document(pdfFile)
    pdfFile.set_parser(parser)
    # 提供初始化密码
    pdfFile.initialize()
    # 检测文档是否提供txt转换
    if not pdfFile.is_extractable:
        raise PDFTextExtractionNotAllowed
    else:
        # 解析数据
        # 数据管理器
        manager = PDFResourceManager()
        # 创建一个PDF设备对象
        laparams = LAParams()
        device = PDFPageAggregator(manager, laparams=laparams)
        # 解释器对象
        interpreter = PDFPageInterpreter(manager, device)
        # 开始循环处理，每次处理一页
        for page in pdfFile.get_pages():
            interpreter.process_page(page)
            #
            layout = device.get_result()
            for x in layout:
                if isinstance(x, LTTextBoxHorizontal):
                    if topath == "":
                        # 处理每行数据
                        str1 = x.get_text()
                        if callback != None:
                            # callback为函数，用于在外部处理数据而不改变readPDF函数
                            callback(str1)
                        else:
                            print(str1)
                    else:
                        with open(topath, "w") as f:
                            str1 = x.get_text()
                            f.write(str1 + "\n")


path = "D:\学习\python\project\pdf1.pdf"
topath = "D:\学习\python\project\pdf2.txt"


# 这种函数叫做回调函数(在另一个函数里执行此函数)
def func(str):
    print(str + "!")


readPDF(path, func)  # WPS2019在读取pdf并写入至其他文件时会损坏pdf。 垃圾WPS！！！


# 播放音乐
import time
import pygame

# 音乐路径
filePath = "D:\学习\python\project\音乐1.mp3"
# 初始化
pygame.mixer.init()
# 加载音乐
track = pygame.mixer.music.load(filePath)
# 播放
pygame.mixer.music.play()
# 播放10秒
time.sleep(10)
# 暂停
pygame.mixer.music.pause()
# 停止
pygame.mixer.music.stop()

"""
# 修改背景图片
# win键+r --> regedit --> HKEY_CURRENT_USER --> Control Panel --> Desktop --> WallPaper
import win32api
import win32con
import win32gui


def setWallPaper(path):
    # 打开注册表
    reg_key = win32api.RegOpenKeyEx(win32con.HKEY_CURRENT_USER
                                    , "Control Panel\\Desktop", 0, win32con.KEY_SET_VALUE)
    # 2 拉伸 0 居中  6 适应   10 填充
    win32api.RegSetValueEx(reg_key, "WallpaperStyle", 0, win32con.REG_SZ, "2")
    # win32api.RegSetValueEx(reg_key, "WallPaper")
    # SPIF_SENDWININICHANGE立即生效
    win32gui.SystemParametersInfo(win32con.SPI_SETDESKWALLPAPER, path, win32con.SPIF_SENDWININICHANGE)


setWallPaper("D:\学习\python\project\图片1.jpg")
"""

# 整蛊小程序
import threading  # 线程

"""  # 边换壁纸边换音乐 
def go():
    pygame.mixer.init()
    while True:
        for i in range(5):
            filePath2 = r"D:\学习\python\project\音乐" + (str(i+1))+".mp3"
            pygame.mixer.music.load(filePath2)
            pygame.mixer.music.play()
            time.sleep(10)
            pygame.mixer.music.stop()


th = threading.Thread(target=go, name="LoopThread")
th.start()
while True:
    for i in range(3):
        filePath3 = r"D:\学习\python\project\图片" + (str(i+1))+".jpg"
        setWallPaper(filePath3)
        time.sleep(10)
"""


# 键盘模拟
"""
import win32con
import win32api
import  time

win32api.keybd_event(91, 0, 0, 0)  # 按键盘
time.sleep(0.1)
win32api.keybd_event(91, 0, win32con.KEYEVENTF_KEYUP, 0)  # 抬起键盘
"""

""" 不断切换
while True:
    win32api.keybd_event(91, 0, 0, 0)
    time.sleep(0.1)
    win32api.keybd_event(77, 0, 0, 0)
    time.sleep(0.1)
    win32api.keybd_event(77, 0, win32con.KEYEVENTF_KEYUP, 0)
    win32api.keybd_event(91, 0, win32con.KEYEVENTF_KEYUP, 0)
    time.sleep(3)
"""

# 鼠标模拟
""" 双击桌面左上第一个文件
import win32con
import win32api
import time

# 设置鼠标的位置
win32api.SetCursorPos([30, 40])
time.sleep(0.1)
# 鼠标左键按下
win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, 0, 0, 0)
# 鼠标左键抬起
win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, 0, 0, 0)
win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, 0, 0, 0)
win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, 0, 0, 0)
"""


# word自动化办公
# 读取doc与docx文件并写入其他文件
import win32com
import win32com.client


def readWordFile(path, topath):
    # 调用系统word功能，可以处理doc和docx两种文件,但是处理WPS docx时会乱码
    mw = win32com.client.Dispatch("Word.Application")
    # 打开文件
    doc = mw.Documents.Open(path)

    for paragraph in doc.Paragraphs:
        line = paragraph.Range.Text
        print(line)

    # 将word的数据保存到另一个文件
    doc.SaveAs(topath, 2)  # 2表示为txt文件
    # 关闭文件
    doc.Close()
    # 退出word
    mw.Quit()


path2 = "D:\学习\python\project\word2.doc"
topath2 = "D:\学习\python\project\word4.txt"
readWordFile(path2, topath2)

# 创建word文件
"""因为本计算机word没有激活因此会报错
import os


def makeWordFile(path, name):
    word = win32com.client.Dispatch("Word Application")
    # 让文档可见
    word.Visible = True
    # 创建文档
    doc = word.Documents.Add()
    # 写内容
    # 从头开始写
    r = doc.Range(0, 0)
    r.InsertAfter("亲爱的" + name + "\n")
    r.InsertAfter("我想尼玛呢\n")
    # 存储文件
    doc.SaveAs(path)
    # 关闭
    doc.Close()
    word.Quit()


names = ["张三"]
for name in names:
    path = os.path.join("D:\学习\python\project", name)
    makeWordFile(path, name)
"""


# 读取xlsx文件
# xlsx xls
# openpyxl -> xlsx
from openpyxl.reader.excel import load_workbook


def readXlsxFile(path):
    dic1 = {}
    # 打开文件
    file = load_workbook(filename=path)
    # 所有表格的名称
    sheets = file.sheetnames
    for sheetName in sheets:
        # 拿出一个表格
        sheet = file[sheetName]
        sheetInfo = []
        # 最大行数，列数，表名
        print(sheet.max_row)
        print(sheet.max_column)
        print(sheet.title)
        # 读取整张表
        for lineNum in range(1, sheet.max_row + 1):
            lineList = []
            for columnNum in range(1, sheet.max_column + 1):
                value = sheet.cell(row=lineNum, column=columnNum).value
                lineList.append(value)
            sheetInfo.append(lineList)
        # 将一张表的数据存到字典里
        dic1[sheetName] = sheetInfo
    return dic1


# 不能处理xls文件
path3 = r"D:\学习\python\project\xlsx1.xlsx"
dic = readXlsxFile(path3)
print(dic)


# 返回xls和xlsx文件内容
# 读取虽然方便，但是这种方式只能写入xls
from collections import OrderedDict    # 有序字典
from pyexcel_xls import get_data    # 读取数据


def readXlsAndXlsxFile(path):
    dic1 = OrderedDict()
    # 抓取数据
    xdata = get_data(path)
    for sheet in xdata:
        dic1[sheet] = xdata[sheet]
    return dic1


dic2 = readXlsAndXlsxFile(path3)
print(dic2)
print(len(dic2))


# 写入xls文件
from pyexcel_xls import save_data


def makeExcelFile(path, data):
    dic1 = OrderedDict()
    for sheetName, sheetValue in data.items():
        dic2 = {}
        dic2[sheetName] = sheetValue
        dic1.update(dic2)
    save_data(path, dic1)


path4 = r"D:\学习\python\project\xlsx3.xls"
makeExcelFile(path4, {"表1": [[1, 2, 3], [4, 5, 6], [7, 8, 9]],
                      "表2": [[11, 22, 33], [44, 55, 66], [77, 88, 99]]})


# 写ppt
"""
import win32com
import win32com.client
"""


def makePPT(path):
    ppt = win32com.client.Dispatch("PowerPoint.Application")
    ppt.Visible = True
    # 增加一个文件
    pptFile = ppt.Presentations.Add()
    # 创建页
    page1 = pptFile.Slides.Add(1, 1)    # 参数1为页数(从1开始)  参数2为类型(根据模板)
    t1 = page1.Shapes[0].TextFrame.TextRange
    t1.Text = "sunck"
    t2 = page1.Shapes[1].TextFrame.TextRange
    t2.Text = "sunck is a good man"

    page2 = pptFile.Slides.Add(2, 2)
    t3 = page2.Shapes[0].TextFrame.TextRange
    t3.Text = "kaige"
    t4 = page2.Shapes[1].TextFrame.TextRange
    t4.Text = "kaige is a good man"
    # 保存
    pptFile.SaveAs(path)  # 无法保存WPS2019PPT??
    pptFile.Close()
    ppt.Quit()


path5 = r"D:\学习\python\project\ppt1.ppt"
makePPT(path)
